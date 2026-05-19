
import openpyxl
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

class TemplateExporter:
    def __init__(self, template_path):
        self.template_path = template_path

    def _find_row(self, ws, value, col=1, start_row=1):
        for r in range(start_row, 1000):
            val = ws.cell(row=r, column=col).value
            if val and isinstance(val, str) and value in val: return r
        return None

    def generate_master_calculation(self, proc, adj_usage, depr_mapping, groups, proj_data, contract_mapping, output_path):
        wb = openpyxl.load_workbook(self.template_path)
        total_hours = sum(adj_usage.values())
        instruments = sorted(adj_usage.keys())
        shares = {inst: (adj_usage[inst] / total_hours if total_hours > 0 else 0) for inst in instruments}
        blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        # --- 1. EXPENDITURES ---
        if 'Expenditures' in wb.sheetnames:
            ws = wb['Expenditures']
            if len(instruments) > 6: ws.insert_cols(10, amount=len(instruments)-6)
            last_col_l = get_column_letter(9 + len(instruments) - 1)
            for i, inst in enumerate(instruments):
                col_idx = 9 + i
                ws.cell(row=9, column=col_idx, value=inst).font = header_font
                ws.cell(row=9, column=col_idx).fill = blue_fill
                ws.cell(row=5, column=col_idx, value=shares[inst])
            for r in range(10, 28):
                for c in range(1, 35): ws.cell(row=r, column=c, value=None)
            current_row = 10
            for i, (_, row) in enumerate(proc.expenses_summary.iterrows()):
                if current_row > 26: break
                ws.cell(row=current_row, column=1, value=row['Financial Account Code'])
                ws.cell(row=current_row, column=2, value=row['Financial Account Title'])
                ws.cell(row=current_row, column=3, value=row['Expense Amount'])
                for j in range(len(instruments)):
                    cl = get_column_letter(9+j)
                    ws.cell(row=current_row, column=9+j, value=f"=$C{current_row}*{cl}$5")
                ws.cell(row=current_row, column=9+len(instruments), value=f"=SUM(I{current_row}:{last_col_l}{current_row})")
                current_row += 1
            contract_row = 27
            ws.cell(row=contract_row, column=2, value="Service Contracts (Allocated)")
            total_contracts = sum([row['Amount'] for _, row in proc.service_contracts.iterrows() if contract_mapping.get(f"{row['Account']}_{row['Description']}_{row['Amount']}", ["Shared"]) != ["Exclude"]])
            ws.cell(row=contract_row, column=3, value=total_contracts)
            sum_across = 0
            for j, inst in enumerate(instruments):
                share = 0
                for _, row in proc.service_contracts.iterrows():
                    key = f"{row['Account']}_{row['Description']}_{row['Amount']}"
                    targets = contract_mapping.get(key, ["Shared"])
                    if targets == ["Shared"]: targets = instruments
                    if inst in targets:
                        gh = sum([adj_usage[t] for t in targets])
                        if gh > 0: share += (adj_usage[inst] / gh) * row['Amount']
                ws.cell(row=contract_row, column=9+j, value=share)
                sum_across += share
            ws.cell(row=contract_row, column=4, value=total_contracts - sum_across)
            ws.cell(row=contract_row, column=9+len(instruments), value=f"=SUM(I{contract_row}:{last_col_l}{contract_row})")
            for j in range(len(instruments) + 1):
                col_l = get_column_letter(9 + j)
                ws.cell(row=28, column=9+j, value=f"=SUM({col_l}10:{col_l}27)")
                ws.cell(row=33, column=9+j, value=f"=SUM({col_l}29:{col_l}32)")
                ws.cell(row=35, column=9+j, value=f"={col_l}28+{col_l}33")
            ws.cell(row=35, column=8, value=f"=H28+H33")

        # --- 2. SALARIES AND WAGES ---
        if 'Salaries and Wages' in wb.sheetnames:
            ws = wb['Salaries and Wages']
            if len(instruments) > 6: ws.insert_cols(11, amount=len(instruments)-6)
            last_col_l = get_column_letter(10 + len(instruments) - 1)
            for i, inst in enumerate(instruments):
                ws.cell(row=11, column=10+i, value=inst).font = header_font
                ws.cell(row=11, column=10+i).fill = blue_fill
                ws.cell(row=5, column=10+i, value=shares[inst])
            for r in range(12, 21):
                for c in range(1, 35): ws.cell(row=r, column=c, value=None)
            for i, (_, row) in enumerate(proc.salaries_summary.iterrows()):
                row_idx = 12 + i
                if row_idx > 20: break
                ws.cell(row=row_idx, column=1, value=row['Employee Name'])
                ws.cell(row=row_idx, column=2, value=row['Position Title'])
                ws.cell(row=row_idx, column=5, value=row['3E Fund Expense Amount']) 
                ws.cell(row=row_idx, column=8, value=1.0)
                ws.cell(row=row_idx, column=9, value=f"=$E{row_idx}*H{row_idx}")
                for j in range(len(instruments)):
                    cl = get_column_letter(10+j)
                    ws.cell(row=row_idx, column=10+j, value=f"=$I{row_idx}*{cl}$5")
                ws.cell(row=row_idx, column=10+len(instruments), value=f"=SUM(J{row_idx}:{last_col_l}{row_idx})")
            for c_idx in [9] + list(range(10, 10 + len(instruments) + 1)):
                col_l = get_column_letter(c_idx)
                ws.cell(row=22, column=c_idx, value=f"=SUM({col_l}12:{col_l}21)")

        # --- 3. EQUIPMENT DEPRECIATION ---
        if 'Equipment Depreciation' in wb.sheetnames:
            ws = wb['Equipment Depreciation']
            start_c = 7 # G
            if len(instruments) > 6: ws.insert_cols(start_c + 1, amount=len(instruments)-6)
            for i, inst in enumerate(instruments):
                ws.cell(row=10, column=start_c+i, value=inst).font = header_font
                ws.cell(row=10, column=start_c+i).fill = blue_fill
            df_3e = proc.dep_full[proc.dep_full['Fund Type'] == '3E']
            sub_3e_row = self._find_row(ws, "Total 3E Equipment Depreciation", col=3, start_row=11) or 17
            if len(df_3e) > (sub_3e_row - 11):
                ws.insert_rows(sub_3e_row, amount=len(df_3e)-(sub_3e_row-11))
                sub_3e_row = 11 + len(df_3e)
            for i, (_, row) in enumerate(df_3e.iterrows()):
                r_idx = 11 + i
                ws.cell(row=r_idx, column=2, value=row['Permanent Tag Ptag'])
                ws.cell(row=r_idx, column=3, value=row['Asset Description'])
                ws.cell(row=r_idx, column=4, value=row['Depreciation Amount'])
                target = depr_mapping.get(row['Asset Description'], "Skip")
                if target in instruments: ws.cell(row=r_idx, column=start_c + instruments.index(target), value=row['Depreciation Amount'])
                ws.cell(row=r_idx, column=start_c + len(instruments), value=f"=SUM({get_column_letter(start_c)}{r_idx}:{get_column_letter(start_c+len(instruments)-1)}{r_idx})")
            for c_idx in [4, 5, 6] + list(range(start_c, start_c + len(instruments) + 1)):
                col_l = get_column_letter(c_idx)
                ws.cell(row=sub_3e_row, column=c_idx, value=f"=SUM({col_l}11:{col_l}{sub_3e_row-1})")
            non3e_label = self._find_row(ws, "Non-3E Equipment", start_row=sub_3e_row) or (sub_3e_row + 2)
            st_non3e = non3e_label + 2
            sub_non3e_row = self._find_row(ws, "Total Non 3E Equipment Depreciation", col=3, start_row=st_non3e) or (st_non3e + 6)
            df_non3e = proc.dep_full[proc.dep_full['Fund Type'] == 'Non-3E']
            if len(df_non3e) > (sub_non3e_row - st_non3e):
                ws.insert_rows(sub_non3e_row, amount=len(df_non3e)-(sub_non3e_row-st_non3e))
                sub_non3e_row = st_non3e + len(df_non3e)
            for i, (_, row) in enumerate(df_non3e.iterrows()):
                r_idx = st_non3e + i
                ws.cell(row=r_idx, column=2, value=row['Permanent Tag Ptag']); ws.cell(row=r_idx, column=3, value=row['Asset Description']); ws.cell(row=r_idx, column=4, value=row['Depreciation Amount'])
                target = depr_mapping.get(row['Asset Description'], "Skip")
                if target in instruments: ws.cell(row=r_idx, column=start_c + instruments.index(target), value=row['Depreciation Amount'])
                ws.cell(row=r_idx, column=start_c + len(instruments), value=f"=SUM({get_column_letter(start_c)}{r_idx}:{get_column_letter(start_c+len(instruments)-1)}{r_idx})")
            for c_idx in [4, 5, 6] + list(range(start_c, start_c + len(instruments) + 1)):
                col_l = get_column_letter(c_idx)
                ws.cell(row=sub_non3e_row, column=c_idx, value=f"=SUM({col_l}{st_non3e}:{col_l}{sub_non3e_row-1})")
            proj_label = self._find_row(ws, "Projections", start_row=sub_non3e_row) or (sub_non3e_row + 2)
            st_proj = proj_label + 1
            sub_proj_row = self._find_row(ws, "Total Projected Equipment Depreciation", col=3, start_row=st_proj) or (st_proj + 4)
            if len(proj_data) > (sub_proj_row - st_proj):
                ws.insert_rows(sub_proj_row, amount=len(proj_data)-(sub_proj_row-st_proj))
                sub_proj_row = st_proj + len(proj_data)
            for i, p in enumerate(proj_data):
                r_idx = st_proj + i
                ws.cell(row=r_idx, column=3, value=p['description']); ws.cell(row=r_idx, column=4, value=p['amount'])
                for t in p['allocation']:
                    if t == "Shared":
                        for k, inst in enumerate(instruments): ws.cell(row=r_idx, column=start_c+k, value=p['amount']*shares[inst])
                    elif t in instruments:
                        ws.cell(row=r_idx, column=start_c+instruments.index(t), value=p['amount'])
                ws.cell(row=r_idx, column=start_c + len(instruments), value=f"=SUM({get_column_letter(start_c)}{r_idx}:{get_column_letter(start_c+len(instruments)-1)}{r_idx})")
            for c_idx in [4, 5, 6] + list(range(start_c, start_c + len(instruments) + 1)):
                col_l = get_column_letter(c_idx)
                ws.cell(row=sub_proj_row, column=c_idx, value=f"=SUM({col_l}{st_proj}:{col_l}{sub_proj_row-1})")
            internal_total_row = self._find_row(ws, "Total Depreciation for Internal Rate", col=3, start_row=sub_proj_row) or (sub_proj_row + 2)
            for c_idx in [4, 5, 6] + list(range(start_c, start_c + len(instruments) + 1)):
                col_l = get_column_letter(c_idx)
                ws.cell(row=internal_total_row, column=c_idx, value=f"={get_column_letter(c_idx)}{sub_3e_row}+{get_column_letter(c_idx)}{sub_non3e_row}+{get_column_letter(c_idx)}{sub_proj_row}")

        # --- 4 & 5. RATE SUMMARIES (COL K TOTAL FIX) ---
        for s_name in ['Rate Summary - Internal', 'Rate Summary - External']:
            if s_name in wb.sheetnames:
                ws = wb[s_name]
                # Template has C to J as slots. Col K is 11.
                # If instruments > 8, we insert more.
                if len(instruments) > 8: ws.insert_cols(11, amount=len(instruments)-8)
                
                last_inst_col_idx = 3 + len(instruments) - 1 # Start Col C(3)
                last_inst_l = get_column_letter(last_inst_col_idx)
                total_col_idx = last_inst_col_idx + 1 # This is K if len=8
                total_l = get_column_letter(total_col_idx)
                
                for i, inst in enumerate(instruments):
                    ws.cell(row=8, column=3+i, value=inst).font = Font(bold=True, color="0000FF")
                
                rows = [10, 11, 12, 14, 16, 17, 19] if "Internal" in s_name else [10, 11, 12, 13, 16, 17, 19, 20, 21, 22, 24, 25, 27]
                
                for r in rows:
                    # HORIZONTAL TOTAL: SUM(C:LastInstrument) -> Put in Total Column
                    ws.cell(row=r, column=total_col_idx, value=f"=SUM(C{r}:{last_inst_l}{r})")
                    
                    # LINK DATA
                    for i in range(len(instruments)):
                        curr_c = 3 + i
                        if r == 10: ws.cell(row=r, column=curr_c, value=f"=Expenditures!{get_column_letter(9+i)}35")
                        elif r == 11: ws.cell(row=r, column=curr_c, value=f"='Salaries and Wages'!{get_column_letter(10+i)}22")
                        elif r == 12: ws.cell(row=r, column=curr_c, value=f"='Equipment Depreciation'!{get_column_letter(7+i)}{internal_total_row}")
                        elif r in [19, 27]: ws.cell(row=r, column=curr_c, value=f"=Base!C{9+i}")
                    
                    # PROPAGATE MATH (Total Expenditures, Rates)
                    cl_tot = get_column_letter(total_col_idx)
                    for i in range(len(instruments)):
                        cl = get_column_letter(3 + i)
                        if "Internal" in s_name:
                            if r == 14: ws.cell(row=r, column=3+i, value=f"=SUM({cl}10:{cl}13)")
                            elif r == 17: ws.cell(row=r, column=3+i, value=f"={cl}14+{cl}16")
                            elif r == 21: ws.cell(row=r, column=3+i, value=f"=IFERROR({cl}17/{cl}19,0)")
                        else:
                            if r == 13: ws.cell(row=r, column=3+i, value=f"=SUM({cl}10:{cl}12)")
                            elif r == 22: ws.cell(row=r, column=3+i, value=f"=SUM({cl}16:{cl}21)")
                            elif r == 25: ws.cell(row=r, column=3+i, value=f"={cl}13+{cl}22+{cl}24")
                            elif r == 29: ws.cell(row=r, column=3+i, value=f"=IFERROR({cl}25/{cl}27,0)")
                            elif r == 33: ws.cell(row=r, column=3+i, value=f"=+{cl}29+({cl}29*{cl}31)")

                # Also update the SUM formulas for the TOTAL columns (B and K)
                ws.cell(row=14, column=total_col_idx, value=f"=SUM({total_l}10:{total_l}13)")
                ws.cell(row=17, column=total_col_idx, value=f"={total_l}14+{total_l}16")
                # Summary B14/B17 for Internal
                if "Internal" in s_name:
                    ws.cell(row=14, column=2, value=f"=SUM(B10:B13)")
                    ws.cell(row=17, column=2, value=f"=B14+B16")

        # (Base/Depr Detail same)
        if 'Base' in wb.sheetnames:
            ws = wb['Base']
            for r in range(9, 60):
                for c in range(1, 5): ws.cell(row=r, column=c, value=None)
            for i, inst in enumerate(instruments):
                ws.cell(row=9+i, column=1, value=inst); ws.cell(row=9+i, column=2, value="Hours"); ws.cell(row=9+i, column=3, value=adj_usage[inst])
            ws.cell(row=16, column=3, value=f"=SUM(C9:C15)")
        if 'Depreciation Detail' in wb.sheetnames:
            ws = wb['Depreciation Detail']
            df_3e = proc.dep_full[proc.dep_full['Fund Type'] == '3E']; df_non3e = proc.dep_full[proc.dep_full['Fund Type'] == 'Non-3E']
            sub_3e = self._find_row(ws, "Total Depreciation", col=8, start_row=11) or 19
            if len(df_3e) > (sub_3e - 11): ws.insert_rows(sub_3e, amount=len(df_3e)-(sub_3e-11)); sub_3e = 11 + len(df_3e)
            for i, (_, row) in enumerate(df_3e.iterrows()):
                for col_idx, value in enumerate(row.values[:-1], 1): ws.cell(row=11+i, column=col_idx, value=value)
            non3e_h = self._find_row(ws, "Non-3E Equipment", start_row=sub_3e) or 21
            st_n3e = non3e_h + 2; sub_n3e = self._find_row(ws, "Total Depreciation", col=8, start_row=st_n3e) or (st_n3e + 8)
            if len(df_non3e) > (sub_n3e - st_n3e): ws.insert_rows(sub_n3e, amount=len(df_non3e)-(sub_n3e-st_n3e))
            for i, (_, row) in enumerate(df_non3e.iterrows()):
                for col_idx, value in enumerate(row.values[:-1], 1): ws.cell(row=st_n3e+i, column=col_idx, value=value)
        if 'Base Detail ' in wb.sheetnames:
            ws = wb['Base Detail ']
            for r in range(8, 200):
                for c in range(1, 30): ws.cell(row=r, column=c, value=None)

        wb.save(output_path)
        return output_path
